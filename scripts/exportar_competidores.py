"""Exporta la lista de competidores de Higiofi a un Excel.

Lee `adjudicaciones` (poblada por scripts/extraer_adjudicaciones.py), filtra
por CPVs del catálogo Higiofi (config.CPVS_RELEVANTES) y genera un Excel con
tres hojas:

1. "Competidores Andalucía" — empresas que han ganado licitaciones de
   CPVs Higiofi en Andalucía (los más relevantes para Cádiz).
2. "Competidores España" — todas las empresas que han ganado CPVs Higiofi
   en cualquier provincia española (visión nacional del mercado).
3. "Adjudicaciones detalle" — una fila por adjudicación con licitación
   completa (uuid, objeto, organismo, fecha, importe, CPV).

Uso:
    python scripts/exportar_competidores.py
    python scripts/exportar_competidores.py --salida ruta/al/competidores.xlsx
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from src.config import CPVS_RELEVANTES, DB_PATH  # noqa: E402
from src.logging_setup import configurar_logging  # noqa: E402
from src.mapa.config_mapa import (  # noqa: E402
    CPVS_PARQUES, cargar_blacklist_competencia,
)


def _categoria_higiofi(cpvs_ganados: str | None) -> str:
    """Devuelve 'Parques' o 'Mobiliario / papelería' según los CPVs ganados.
    Si gana en ambos, prevalece Parques (más especializado)."""
    if not cpvs_ganados:
        return "Mobiliario / papelería"
    for cpv in cpvs_ganados.split(","):
        cpv = cpv.strip()
        if any(cpv.startswith(p) for p in CPVS_PARQUES):
            return "Parques"
    return "Mobiliario / papelería"


PROVINCIAS_ANDALUCIA = (
    "Cádiz", "Sevilla", "Málaga", "Granada", "Huelva", "Jaén", "Córdoba", "Almería",
)

# Heurística: keywords en razón social que sugieren "constructora" más que
# "proveedor de mobiliario". Una constructora gana licitaciones de obra que
# incluyen parques/mobiliario pero COMPRA el material a proveedores como
# Higiofi — es cliente potencial, no competencia.
_KEYWORDS_CONSTRUCTORA = (
    "CONSTRUC", "OBRA", "EDIFIC", "INGENIER",
    "RESTAURACION", "REFORMAS", "PROMOCIONES",
)


def _agregado_por_nif(conn: sqlite3.Connection, solo_andalucia: bool) -> list[dict]:
    """Competencia: empresas que ganaron CPV Higiofi como SUMINISTRO (no obra).

    Excluye razones sociales que suenan a constructora (las constructoras son
    clientes potenciales, no competencia).
    """
    cpv_or = " OR ".join([f"l.cpv_principal LIKE '{p}%'" for p in CPVS_RELEVANTES])
    where_provincia = ""
    params: list = []
    if solo_andalucia:
        placeholders = ",".join("?" * len(PROVINCIAS_ANDALUCIA))
        where_provincia = f"AND l.provincia IN ({placeholders})"
        params.extend(PROVINCIAS_ANDALUCIA)
    name_excl = " AND ".join([f"UPPER(a.razon_social) NOT LIKE '%{kw}%'" for kw in _KEYWORDS_CONSTRUCTORA])

    q = f"""
        SELECT
            a.nif,
            MAX(a.razon_social)            AS razon_social,
            MAX(a.ciudad)                  AS ciudad,
            MAX(a.provincia)               AS provincia,
            MAX(a.codigo_postal)           AS cp,
            MAX(a.es_pyme)                 AS es_pyme,
            COUNT(*)                       AS n_adjudicaciones,
            SUM(a.importe_adjudicacion)    AS importe_total,
            MIN(a.fecha_adjudicacion)      AS primera_fecha,
            MAX(a.fecha_adjudicacion)      AS ultima_fecha,
            GROUP_CONCAT(DISTINCT l.cpv_principal) AS cpvs_ganados,
            GROUP_CONCAT(DISTINCT l.provincia)     AS provincias_ganadas
        FROM adjudicaciones a
        JOIN licitaciones l ON a.uuid_placsp = l.uuid_placsp
        WHERE ({cpv_or})
          AND l.tipo_contrato = 'Suministros'
          AND a.nif IS NOT NULL
          AND a.razon_social IS NOT NULL
          AND {name_excl}
          {where_provincia}
        GROUP BY a.nif
        ORDER BY n_adjudicaciones DESC, importe_total DESC
    """
    return [dict(r) for r in conn.execute(q, params)]


def _adjudicaciones_detalle(conn: sqlite3.Connection) -> list[dict]:
    """Una fila por adjudicación con info de licitación cruzada."""
    cpv_or = " OR ".join([f"l.cpv_principal LIKE '{p}%'" for p in CPVS_RELEVANTES])
    q = f"""
        SELECT
            a.nif, a.razon_social, a.ciudad AS empresa_ciudad,
            a.provincia AS empresa_provincia,
            a.fecha_adjudicacion, a.importe_adjudicacion,
            a.es_pyme,
            l.uuid_placsp, l.expediente, l.objeto, l.cpv_principal,
            l.organo_contratacion, l.municipio AS licitacion_municipio,
            l.provincia AS licitacion_provincia,
            l.enlace_placsp
        FROM adjudicaciones a
        JOIN licitaciones l ON a.uuid_placsp = l.uuid_placsp
        WHERE ({cpv_or})
          AND a.nif IS NOT NULL
          AND a.razon_social IS NOT NULL
        ORDER BY a.fecha_adjudicacion DESC, a.razon_social
    """
    return [dict(r) for r in conn.execute(q)]


# --- Helpers de formato Excel ----------------------------------------------
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def _escribir_hoja(ws, columnas: list[tuple[str, str, int]], filas: list[dict]) -> None:
    """Escribe una hoja con cabecera estilizada, datos y autofiltro.

    columnas: lista de (clave_dict, etiqueta, ancho).
    """
    for i, (_, etiqueta, ancho) in enumerate(columnas, start=1):
        c = ws.cell(row=1, column=i, value=etiqueta)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, (clave, _, _) in enumerate(columnas, start=1):
            valor = fila.get(clave)
            # Convertir booleanos / 0/1 PYME a Sí/No
            if clave == "es_pyme":
                if valor == 1:
                    valor = "Sí"
                elif valor == 0:
                    valor = "No"
                else:
                    valor = ""
            ws.cell(row=row_idx, column=col_idx, value=valor)

    # Autofiltro
    if filas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{len(filas)+1}"

    # Pin de la primera fila
    ws.freeze_panes = "A2"


COLUMNAS_RESUMEN: list[tuple[str, str, int]] = [
    ("categoria",        "Categoría",           22),
    ("razon_social",     "Razón social",        45),
    ("nif",              "NIF",                 14),
    ("ciudad",           "Ciudad sede",         22),
    ("provincia",        "Provincia sede",      18),
    ("cp",               "CP",                  8),
    ("n_adjudicaciones", "Nº adjudicaciones",   18),
    ("importe_total",    "Importe total (€)",   18),
    ("primera_fecha",    "Primera victoria",    14),
    ("ultima_fecha",     "Última victoria",     14),
    ("es_pyme",          "PYME",                8),
    ("provincias_ganadas", "Provincias donde gana",  28),
    ("cpvs_ganados",     "CPVs ganados",        30),
]

COLUMNAS_DETALLE: list[tuple[str, str, int]] = [
    ("fecha_adjudicacion",  "Fecha",              12),
    ("razon_social",        "Adjudicatario",      40),
    ("nif",                 "NIF",                14),
    ("empresa_ciudad",      "Ciudad empresa",     20),
    ("empresa_provincia",   "Prov. empresa",      14),
    ("importe_adjudicacion","Importe (€)",        14),
    ("cpv_principal",       "CPV",                12),
    ("objeto",              "Objeto",             60),
    ("organo_contratacion", "Órgano contratante", 40),
    ("licitacion_municipio","Municipio licit.",   22),
    ("licitacion_provincia","Provincia licit.",   14),
    ("expediente",          "Expediente",         18),
    ("enlace_placsp",       "Enlace PLACSP",      30),
]


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--salida", default="data/competidores_higiofi.xlsx",
                   help="Ruta del Excel de salida (default: data/competidores_higiofi.xlsx)")
    args = p.parse_args()

    log = configurar_logging("exportar_competidores")
    salida = Path(args.salida)
    if not salida.is_absolute():
        salida = (Path(__file__).resolve().parent.parent / salida).resolve()
    salida.parent.mkdir(parents=True, exist_ok=True)

    blacklist = cargar_blacklist_competencia()
    if blacklist:
        log.info("Blacklist activa: %d NIFs excluidos", len(blacklist))

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        log.info("Consultando adjudicaciones...")
        andalucia = [e for e in _agregado_por_nif(conn, solo_andalucia=True) if e["nif"] not in blacklist]
        espana = [e for e in _agregado_por_nif(conn, solo_andalucia=False) if e["nif"] not in blacklist]
        # Añade columna "categoría" (Parques / Mobiliario) a cada empresa
        for fila in andalucia + espana:
            fila["categoria"] = _categoria_higiofi(fila.get("cpvs_ganados"))
        detalle = _adjudicaciones_detalle(conn)
    finally:
        conn.close()

    log.info("Competidores en Andalucía: %d", len(andalucia))
    log.info("Competidores en España:    %d", len(espana))
    log.info("Adjudicaciones detalle:    %d", len(detalle))

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Competidores Andalucía")
    _escribir_hoja(ws1, COLUMNAS_RESUMEN, andalucia)

    ws2 = wb.create_sheet("Competidores España")
    _escribir_hoja(ws2, COLUMNAS_RESUMEN, espana)

    ws3 = wb.create_sheet("Adjudicaciones detalle")
    _escribir_hoja(ws3, COLUMNAS_DETALLE, detalle)

    wb.save(str(salida))
    log.info("Excel escrito: %s", salida)
    log.info("Tamaño: %.1f KB", salida.stat().st_size / 1024)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
